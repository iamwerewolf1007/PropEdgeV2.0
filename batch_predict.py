"""
PropEdge V2.0 — batch_predict.py
Fetches today's props from Odds API, computes rolling features from game logs,
runs V2.0 LightGBM inference, generates narratives, appends props to Excel,
writes data/today.json, pushes to GitHub.

RUN: python3 run.py predict
     python3 run.py fetch-and-predict
"""
from __future__ import annotations

import json
import time
import warnings
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd

from config import (
    SOURCE_DIR, DATA_DIR, FILE_GL_2526, FILE_GL_2425,
    FILE_H2H, FILE_PROPS_2526, FILE_TODAY, FILE_DVP,
    ODDS_API_KEY, ODDS_API_BASE, ODDS_SPORT, ODDS_MARKET, CREDIT_ALERT,
    VERSION_TAG, MIN_PRIOR_GAMES, assign_tier, clean_json, today_et,
    TIER_STAKES, now_et, current_season,
)
from player_name_aliases import _norm, resolve_name, build_nmap
from rolling_engine import (
    filter_played, build_player_index, get_prior_games,
    build_dynamic_dvp, build_pace_rank, build_opp_def_caches,
    build_rest_days_map, extract_features,
)
from reasoning_engine import generate_pre_match_reason

warnings.filterwarnings("ignore")
ET = ZoneInfo("America/New_York")

# ── Team abbreviation map ─────────────────────────────────────────────────────
_TEAM_ABR: dict[str, str] = {
    "Atlanta Hawks": "ATL", "Boston Celtics": "BOS", "Brooklyn Nets": "BKN",
    "Charlotte Hornets": "CHA", "Chicago Bulls": "CHI", "Cleveland Cavaliers": "CLE",
    "Dallas Mavericks": "DAL", "Denver Nuggets": "DEN", "Detroit Pistons": "DET",
    "Golden State Warriors": "GSW", "Houston Rockets": "HOU", "Indiana Pacers": "IND",
    "LA Clippers": "LAC", "Los Angeles Clippers": "LAC", "Los Angeles Lakers": "LAL",
    "LA Lakers": "LAL", "Memphis Grizzlies": "MEM", "Miami Heat": "MIA",
    "Milwaukee Bucks": "MIL", "Minnesota Timberwolves": "MIN", "New Orleans Pelicans": "NOP",
    "New York Knicks": "NYK", "Oklahoma City Thunder": "OKC", "Orlando Magic": "ORL",
    "Philadelphia 76ers": "PHI", "Phoenix Suns": "PHX", "Portland Trail Blazers": "POR",
    "Sacramento Kings": "SAC", "San Antonio Spurs": "SAS", "Toronto Raptors": "TOR",
    "Utah Jazz": "UTA", "Washington Wizards": "WAS",
}

def _abr(t): return _TEAM_ABR.get(t, t[:3].upper())


# ── Odds API fetch ────────────────────────────────────────────────────────────

def _api_get(url: str, params: dict, timeout: int = 30):
    try:
        import requests as _r
        resp = _r.get(url, params=params, timeout=timeout)
        resp.raise_for_status()
        return resp.json(), dict(resp.headers)
    except ImportError:
        import urllib.request, urllib.parse, ssl, json as _j
        qs  = urllib.parse.urlencode(params)
        ctx = ssl.create_default_context()
        try:
            import certifi
            ctx = ssl.create_default_context(cafile=certifi.where())
        except ImportError:
            pass
        with urllib.request.urlopen(f"{url}?{qs}", timeout=timeout, context=ctx) as r:
            return _j.loads(r.read()), dict(r.headers)


def _credits(headers: dict, label: str = "") -> None:
    n = headers.get("x-requests-remaining", headers.get("X-Requests-Remaining", "?"))
    try:
        ni = int(n)
        suffix = " ⚠ LOW" if ni <= CREDIT_ALERT else ""
        print(f"    Credits: {ni}{suffix}  {label}")
    except (ValueError, TypeError):
        pass


def fetch_props_from_api(date_str: str) -> dict[str, dict]:
    """
    Fetch NBA player points props from Odds API.
    Returns {event_id: {home, away, home_raw, away_raw, gametime, props: {player: {...}}}}
    """
    d   = datetime.strptime(date_str, "%Y-%m-%d")
    frm = (d - timedelta(hours=6)).strftime("%Y-%m-%dT%H:%M:%SZ")
    to  = (d + timedelta(hours=30)).strftime("%Y-%m-%dT%H:%M:%SZ")

    print(f"\n  Fetching events: {date_str}")
    body, hdrs = _api_get(
        f"{ODDS_API_BASE}/sports/{ODDS_SPORT}/events",
        {"apiKey": ODDS_API_KEY, "dateFormat": "iso",
         "commenceTimeFrom": frm, "commenceTimeTo": to},
    )
    _credits(hdrs, "events")

    events = []
    for e in body:
        ts = e.get("commence_time", "")
        try:
            utc_dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            et_dt  = utc_dt.astimezone(ET)
            if et_dt.strftime("%Y-%m-%d") == date_str:
                e["_gametime"] = et_dt.strftime("%-I:%M %p ET")
                events.append(e)
        except Exception:
            continue
    print(f"    {len(events)} games on {date_str} ET")

    games: dict[str, dict] = {}
    for e in events:
        eid = e["id"]
        ht  = _abr(e.get("home_team", ""))
        at  = _abr(e.get("away_team", ""))
        games[eid] = {
            "home": ht, "away": at,
            "home_raw": e.get("home_team", ""),
            "away_raw":  e.get("away_team", ""),
            "gametime":  e.get("_gametime", ""),
            "commence":  e.get("commence_time", ""),
            "props":     {},
            "_raw_lines": {},
        }

    for eid, g in games.items():
        time.sleep(0.3)
        try:
            body2, hdrs2 = _api_get(
                f"{ODDS_API_BASE}/sports/{ODDS_SPORT}/events/{eid}/odds",
                {"apiKey": ODDS_API_KEY, "regions": "us",
                 "markets": f"{ODDS_MARKET},spreads,totals",
                 "oddsFormat": "american", "dateFormat": "iso"},
            )
            _credits(hdrs2)
            for bm in body2.get("bookmakers", []):
                for mkt in bm.get("markets", []):
                    if mkt.get("key") != ODDS_MARKET:
                        continue
                    for o in mkt.get("outcomes", []):
                        player = (o.get("description") or "").strip() or o.get("name", "").strip()
                        pt     = o.get("point")
                        side   = (o.get("name") or "").upper()
                        price  = o.get("price")
                        if not player or pt is None:
                            continue
                        g["_raw_lines"].setdefault(player, [])
                        g["_raw_lines"][player].append(float(pt))
                        if player not in g["props"]:
                            g["props"][player] = {"line": float(pt), "over": None, "under": None,
                                                   "books": 0, "min_line": float(pt), "max_line": float(pt)}
                        pd_ = g["props"][player]
                        pd_["min_line"] = min(pd_["min_line"], float(pt))
                        pd_["max_line"] = max(pd_["max_line"], float(pt))
                        if side == "OVER":
                            pd_["over"] = int(price) if price else -110
                            pd_["books"] += 1
                        elif side == "UNDER":
                            pd_["under"] = int(price) if price else -110
            print(f"    ✓ {g['away']} @ {g['home']}: {len(g['props'])} props")
        except Exception as ex:
            print(f"    ✗ {g.get('away_raw','')} @ {g.get('home_raw','')}: {ex}")
            time.sleep(1)

    total = sum(len(g["props"]) for g in games.values())
    print(f"  Total: {total} props across {len(games)} games")
    return games


# ── Excel append ──────────────────────────────────────────────────────────────

def append_props_to_excel(games: dict, date_str: str) -> None:
    """Append fetched props to Player_lines_2025-26.xlsx (source file)."""
    rows = []
    for g in games.values():
        ht = g["home"]; at = g["away"]
        matchup = f"{at} @ {ht}"
        for pname, pd_ in g["props"].items():
            if not pd_.get("line"):
                continue
            rows.append({
                "Date":         pd.Timestamp(date_str),
                "Player":       pname,
                "Position":     "",
                "Game":         matchup,
                "Home":         ht,
                "Away":         at,
                "Game_Time_ET": g.get("gametime", ""),
                "Line":         pd_["line"],
                "Over Odds":    pd_.get("over", -110),
                "Under Odds":   pd_.get("under", -110),
                "Books":        pd_.get("books", 1),
                "Min Line":     pd_.get("min_line", pd_["line"]),
                "Max Line":     pd_.get("max_line", pd_["line"]),
                "Commence":     g.get("commence", ""),
            })
    if not rows:
        print("  ⚠ No props to append to Excel")
        return

    new_df = pd.DataFrame(rows)
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)

    if FILE_PROPS_2526.exists():
        try:
            existing = pd.read_excel(FILE_PROPS_2526, sheet_name="Player_Points_Props")
            existing["Date"] = pd.to_datetime(existing["Date"], errors="coerce")
            # Deduplicate: remove old rows for same date (will replace with fresh fetch)
            existing = existing[existing["Date"].dt.strftime("%Y-%m-%d") != date_str]
            combined = pd.concat([existing, new_df], ignore_index=True)
        except Exception:
            combined = new_df
    else:
        combined = new_df

    combined = combined.sort_values(["Date", "Player"]).reset_index(drop=True)

    try:
        from openpyxl import load_workbook
        if FILE_PROPS_2526.exists():
            wb = load_workbook(FILE_PROPS_2526)
        else:
            from openpyxl import Workbook
            wb = Workbook()
            wb.active.title = "Player_Points_Props"

        ws = wb["Player_Points_Props"] if "Player_Points_Props" in wb.sheetnames else wb.active
        # Clear and rewrite
        ws.delete_rows(1, ws.max_row)
        headers = list(combined.columns)
        ws.append(headers)
        for _, row in combined.iterrows():
            ws.append([
                v.isoformat() if hasattr(v, "isoformat") else
                (None if pd.isna(v) else v)
                for v in row.values
            ])
        wb.save(FILE_PROPS_2526)
        print(f"  ✓ Excel: {FILE_PROPS_2526.name} → {len(combined):,} rows (+{len(new_df)} today)")
    except Exception as e:
        print(f"  ⚠ Excel append failed: {e}")


# ── V2.0 model inference ──────────────────────────────────────────────────────

def score_play(f: dict, line: float, over_odds: int, under_odds: int, books: int) -> dict:
    """Run V2.0 LightGBM + isotonic on a feature dict. Returns model output fields."""
    from model_engine import load_model, load_thresholds, assign_tier as _at
    from config import NEVER_USE_AS_FEATURES

    model, cal, feat_cols = load_model()
    thresholds = load_thresholds()

    safe = [c for c in feat_cols if c not in NEVER_USE_AS_FEATURES]
    row  = {c: f.get(c, 0.0) or 0.0 for c in safe}
    X    = pd.DataFrame([row])[safe].fillna(0).values

    raw_p = float(model.predict_proba(X)[0, 1])
    p_over = float(np.clip(cal.predict([raw_p])[0], 0.01, 0.99))
    p_under = 1.0 - p_over

    direction = "OVER" if p_over >= 0.5 else "UNDER"
    dir_conf  = p_over if direction == "OVER" else p_under
    tier      = _at(dir_conf, thresholds)
    conf_pct  = round(dir_conf * 100, 1)

    # Implied probs from American odds
    def american_to_prob(odds):
        o = int(odds) if odds else -110
        return (-o / (-o + 100)) if o < 0 else (100 / (o + 100))

    imp_over  = american_to_prob(over_odds)
    imp_under = american_to_prob(under_odds)
    market_edge_over  = round(p_over  - imp_over,  4)
    market_edge_under = round(p_under - imp_under, 4)

    L10 = f.get("L10", line)
    L30 = f.get("L30", line)
    pred_pts = round(p_over * L10 + p_under * L30, 1)

    lz = f.get("line_vs_l30", 0) / max(f.get("std10", 5), 1)
    value_flag = (
        "strong_over"  if lz <= -0.5 else
        "mod_over"     if lz <= 0    else
        "mod_under"    if lz <= 0.5  else "strong_under"
    )

    return {
        "p_over":             round(p_over, 4),
        "p_under":            round(p_under, 4),
        "direction":          direction,
        "dir_conf":           round(dir_conf, 4),
        "confidence_pct":     conf_pct,
        "conf":               round(dir_conf, 4),
        "tier":               tier,
        "elite_tier":         tier,
        "elite_prob":         round(dir_conf, 4),
        "elite_stake":        TIER_STAKES.get(tier, 0.0),
        "predicted_pts":      pred_pts,
        "predPts":            pred_pts,
        "predGap":            round(pred_pts - line, 2),
        "value_flag":         value_flag,
        "market_edge_over":   market_edge_over,
        "market_edge_under":  market_edge_under,
        "calProb":            round(p_over, 4),
        "v12_clf_prob":       round(p_over, 4),
    }


# ── Main predict pipeline ─────────────────────────────────────────────────────

def run_predict(date_str: str | None = None, push: bool = True, games: dict | None = None) -> None:
    if date_str is None:
        date_str = today_et()

    print(f"\n  {VERSION_TAG} — Predict  [{date_str}]")
    print("  " + "─" * 52)

    # 1. Fetch props (or use pre-fetched games dict)
    if games is None:
        games = fetch_props_from_api(date_str)
    if not games or not any(g["props"] for g in games.values()):
        print("  ✗ No props fetched.")
        return

    # 2. Append to Excel
    append_props_to_excel(games, date_str)

    # 3. Load game logs + build caches
    print("\n  Building feature caches...")
    dfs = []
    for fp in [FILE_GL_2425, FILE_GL_2526]:
        if fp.exists():
            dfs.append(pd.read_csv(fp, parse_dates=["GAME_DATE"], low_memory=False))
    if not dfs:
        print("  ✗ Game logs not found in source-files/")
        return
    combined = pd.concat(dfs, ignore_index=True)
    played   = filter_played(combined)
    pidx     = build_player_index(played)
    nmap     = {_norm(k): k for k in pidx}
    dvp      = build_dynamic_dvp(played)
    pace     = build_pace_rank(played)
    otr, ovr = build_opp_def_caches(played)
    rmap     = build_rest_days_map(played)

    # 4. Load H2H
    h2h_lkp: dict = {}
    if FILE_H2H.exists():
        df_h2h = pd.read_csv(FILE_H2H, low_memory=False)
        h2h_lkp = {
            (_norm(str(r.get("PLAYER_NAME", ""))), str(r.get("OPPONENT", "")).strip().upper()): r.to_dict()
            for _, r in df_h2h.iterrows()
        }

    # 5. Score each play
    season = current_season()
    plays  = []
    skipped = {"low_line": 0, "no_player": 0, "few_games": 0, "no_features": 0}

    for eid, g in games.items():
        ht = g["home"]; at = g["away"]
        matchup = f"{at} @ {ht}"

        for pname_raw, pd_ in g["props"].items():
            line = pd_.get("line")
            if not line or float(line) < 3:
                skipped["low_line"] += 1
                continue

            # Resolve name
            player = resolve_name(pname_raw, nmap)
            if player is None:
                skipped["no_player"] += 1
                continue

            prior = get_prior_games(pidx, player, date_str)
            if len(prior) < MIN_PRIOR_GAMES:
                skipped["few_games"] += 1
                continue

            last = prior.iloc[-1]
            ptm  = str(last.get("GAME_TEAM_ABBREVIATION", "")).upper()
            ih   = ptm == ht
            opp  = at if ih else ht
            pos_raw = str(last.get("PLAYER_POSITION", "G"))
            rd   = rmap.get((player, date_str), 2)
            h2h_row = h2h_lkp.get((_norm(player), opp.upper()))

            f = extract_features(
                prior=prior, line=float(line), opponent=opp,
                rest_days=rd, pos_raw=pos_raw, game_date=pd.Timestamp(date_str),
                min_line=pd_.get("min_line"), max_line=pd_.get("max_line"),
                dyn_dvp=dvp, pace_rank=pace, opp_trend=otr, opp_var=ovr,
                is_home=ih, h2h_row=h2h_row,
            )
            if f is None:
                skipped["no_features"] += 1
                continue

            # Add market features
            f["books_log"]          = float(np.log1p(pd_.get("books", 1)))
            f["implied_over_prob"]  = _american_to_prob(pd_.get("over", -110))
            f["implied_under_prob"] = _american_to_prob(pd_.get("under", -110))
            f["books"]              = pd_.get("books", 1)

            # Score
            model_out = score_play(
                f, float(line),
                pd_.get("over", -110), pd_.get("under", -110), pd_.get("books", 1)
            )

            # Recent 20 games
            pts_vals   = prior["PTS"].values[-20:]
            dates_vals = prior["GAME_DATE"].values[-20:]
            home_vals  = prior["IS_HOME"].values[-20:] if "IS_HOME" in prior.columns else [True] * len(pts_vals)
            hr10 = float((pts_vals > float(line)).mean()) if len(pts_vals) > 0 else 0.5
            hr30 = hr10

            # Build play dict — all fields dashboard expects
            play = {
                "player":       player,
                "date":         date_str,
                "match":        matchup,
                "fullMatch":    matchup,
                "game":         matchup,
                "home":         ht,
                "away":         at,
                "opponent":     opp,
                "position":     pos_raw,
                "isHome":       bool(ih),
                "gameTime":     g.get("gametime", ""),
                "game_time":    g.get("gametime", ""),
                "team":         ptm,
                "ptm":          ptm,
                "season":       season,
                # Line
                "line":         float(line),
                "overOdds":     pd_.get("over", -110),
                "underOdds":    pd_.get("under", -110),
                "books":        pd_.get("books", 1),
                "min_line":     pd_.get("min_line", float(line)),
                "max_line":     pd_.get("max_line", float(line)),
                "lineHistory":  [{"line": float(line), "batch": 1,
                                  "ts": now_et().strftime("%H:%M")}],
                # Rolling
                "l3":   round(f.get("L3",  0), 1),
                "l5":   round(f.get("L5",  0), 1),
                "l10":  round(f.get("L10", 0), 1),
                "l20":  round(f.get("L20", 0), 1),
                "l30":  round(f.get("L30", 0), 1),
                "std10":round(f.get("std10", 0), 1),
                "hr10": round(hr10, 3),
                "hr30": round(hr30, 3),
                # Model
                **model_out,
                "dir":          model_out["direction"],
                "direction":    model_out["direction"],
                "tierLabel":    model_out["tier"],
                "units":        model_out["elite_stake"],
                "flags":        _count_flags(f, model_out["direction"]),
                "flagsStr":     f"{_count_flags(f, model_out['direction'])}/10",
                "flagDetails":  _flag_details(f, model_out["direction"]),
                # Form
                "volume":       round(f.get("volume",   0), 1),
                "trend":        round(f.get("trend",    0), 1),
                "momentum":     round(f.get("momentum", 0), 1),
                # Minutes/usage
                "min_l10":      round(f.get("min_l10", 28), 1),
                "minL10":       round(f.get("min_l10", 28), 1),
                "min_l30":      round(f.get("min_l30", 28), 1),
                "usage_l10":    round(f.get("usage_l10", 0.18), 3),
                "fta_l10":      round(f.get("fta_l10", 0), 1),
                "fga_l10":      round(f.get("fga_l10", 0), 1),
                "fg3a_l10":     round(f.get("fg3a_l10", 0), 1),
                "pts_per_min":  round(f.get("pts_per_min", 0.5), 3),
                "home_away_split": round(f.get("home_away_split", 0), 1),
                "homeAvgPts":   round(f.get("home_l10", f.get("L10", 0)), 1),
                "awayAvgPts":   round(f.get("away_l10", f.get("L10", 0)), 1),
                "level_ewm":    round(f.get("level_ewm", f.get("L10", 0)), 1),
                "line_vs_l30":  round(f.get("line_vs_l30", 0), 2),
                # Context
                "is_b2b":       bool(f.get("is_b2b", 0)),
                "rest_days":    int(rd),
                "defP":         int(f.get("defP_dynamic", 15)),
                "defP_dynamic": int(f.get("defP_dynamic", 15)),
                "pace_rank":    int(f.get("pace_rank", 15)),
                "pace":         int(f.get("pace_rank", 15)),
                "seasonProgress": round(f.get("season_progress", 0.5), 3),
                "meanReversionRisk": round(f.get("mean_reversion_risk", 0), 2),
                "extreme_hot":  bool(f.get("extreme_hot", False)),
                "extreme_cold": bool(f.get("extreme_cold", False)),
                # H2H
                "h2hG":         int(f.get("h2h_games", 0)),
                "h2h_games":    int(f.get("h2h_games", 0)),
                "h2h":          round(f.get("h2h_avg"), 1) if f.get("h2h_avg") is not None else None,
                "h2h_avg":      round(f.get("h2h_avg"), 1) if f.get("h2h_avg") is not None else None,
                "h2hTsDev":     round(f.get("h2h_ts_dev", 0), 2),
                "h2hFgaDev":    round(f.get("h2h_fga_dev", 0), 2),
                "h2hConfidence":round(f.get("h2h_conf", 0), 3),
                # Recent
                "recent20":     [float(v) for v in pts_vals],
                "recent20dates":[str(pd.Timestamp(d).date()) for d in dates_vals],
                "recent20homes":[bool(v) for v in home_vals],
                # Grade fields
                "result":       "",
                "actualPts":    None,
                "delta":        None,
                "lossType":     None,
                "preMatchReason":  "",
                "postMatchReason": "",
                # V2.0 extras
                "all_clf_agree": bool(f.get("all_windows_over", 0) or f.get("all_windows_under", 0)),
                "trust_mean":    0.70,
                "q25_v12":       round(float(line) - f.get("std10", 5), 1),
                "q75_v12":       round(float(line) + f.get("std10", 5), 1),
                "q_confidence":  round(model_out["dir_conf"], 3),
                "reg_consensus": bool(f.get("all_windows_over", 0) or f.get("all_windows_under", 0)),
                "v12_extreme":   bool(model_out["dir_conf"] >= 0.95),
                "real_gap_v92":  round(f.get("L30", 0) - float(line), 2),
                "real_gap_v12":  round(model_out["predGap"], 2),
            }

            # Generate narrative
            play["preMatchReason"] = generate_pre_match_reason(play)
            plays.append(play)

    # Summary
    skipped_total = sum(skipped.values())
    print(f"\n  Scored: {len(plays)} plays | Skipped: {skipped_total}")
    if skipped_total:
        print(f"    ({', '.join(f'{v} {k}' for k,v in skipped.items() if v)})")

    tier_counts: dict[str, int] = {}
    for p in plays:
        t = p.get("tier", "SKIP")
        tier_counts[t] = tier_counts.get(t, 0) + 1
    for tier in ["APEX", "ULTRA", "ELITE", "STRONG", "PLAY", "SKIP"]:
        n = tier_counts.get(tier, 0)
        if n > 0:
            print(f"    {tier:<8} {n:3d}")

    # Merge with existing today.json
    existing: list[dict] = []
    if FILE_TODAY.exists():
        try:
            with open(FILE_TODAY) as f_:
                existing = json.load(f_)
        except Exception:
            pass

    ex_map = {(p.get("player"), p.get("date")): p for p in existing
              if p.get("result") in ("WIN", "LOSS", "DNP", "PUSH")}
    new_map = {(p["player"], p["date"]): p for p in plays}
    merged  = list(ex_map.values()) + [p for p in plays if (p["player"], p["date"]) not in ex_map]
    merged.sort(key=lambda p: (p.get("tier", "Z"), -p.get("conf", 0)))

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(FILE_TODAY, "w") as f_:
        json.dump(clean_json(merged), f_, indent=2)
    print(f"\n  ✓ today.json → {len(merged)} plays")

    if push:
        from git_push import push as git_push
        git_push(f"V2.0 predict {date_str} — {len(plays)} plays", grade=False)


def _american_to_prob(odds) -> float:
    o = int(odds) if odds else -110
    return (-o / (-o + 100)) if o < 0 else (100 / (o + 100))


def _count_flags(f: dict, direction: str) -> int:
    is_over = direction == "OVER"
    score = 0
    if (is_over and f.get("volume", 0) > 0) or (not is_over and f.get("volume", 0) < 0): score += 1
    if (is_over and f.get("hr30", 0.5) > 0.5) or (not is_over and f.get("hr30", 0.5) < 0.5): score += 1
    if (is_over and f.get("hr10", 0.5) > 0.5) or (not is_over and f.get("hr10", 0.5) < 0.5): score += 1
    if (is_over and f.get("trend", 0) > 0) or (not is_over and f.get("trend", 0) < 0): score += 1
    if (is_over and f.get("l10_vs_line", 0) > 0) or (not is_over and f.get("l10_vs_line", 0) < 0): score += 1
    if (is_over and f.get("defP_dynamic", 15) > 15) or (not is_over and f.get("defP_dynamic", 15) < 15): score += 1
    h2h = f.get("h2h_avg")
    if h2h is not None:
        line = f.get("line", 20)
        if (is_over and h2h > line) or (not is_over and h2h < line): score += 1
    if (is_over and f.get("pace_rank", 15) > 15) or (not is_over and f.get("pace_rank", 15) < 15): score += 1
    if f.get("all_windows_over") or f.get("all_windows_under"): score += 1
    if (is_over and not f.get("is_b2b", False)) or (not is_over): score += 1
    return min(score, 10)


def _flag_details(f: dict, direction: str) -> list[dict]:
    is_over = direction == "OVER"
    line = f.get("line", 20)
    h2h_avg = f.get("h2h_avg")
    details = [
        {"name": "Volume",   "agrees": (is_over and f.get("volume",0)>0) or (not is_over and f.get("volume",0)<0),   "detail": f"{f.get('volume',0):+.1f}"},
        {"name": "HR L30",   "agrees": (is_over and f.get("hr30",0.5)>0.5) or (not is_over and f.get("hr30",0.5)<0.5), "detail": f"{f.get('hr30',0.5)*100:.0f}%"},
        {"name": "HR L10",   "agrees": (is_over and f.get("hr10",0.5)>0.5) or (not is_over and f.get("hr10",0.5)<0.5), "detail": f"{f.get('hr10',0.5)*100:.0f}%"},
        {"name": "Trend",    "agrees": (is_over and f.get("trend",0)>0) or (not is_over and f.get("trend",0)<0),     "detail": f"{f.get('trend',0):+.1f}"},
        {"name": "L10vsLine","agrees": (is_over and f.get("l10_vs_line",0)>0) or (not is_over and f.get("l10_vs_line",0)<0), "detail": f"{f.get('l10_vs_line',0):+.1f}"},
        {"name": "Defense",  "agrees": (is_over and f.get("defP_dynamic",15)>15) or (not is_over and f.get("defP_dynamic",15)<15), "detail": f"#{f.get('defP_dynamic',15):.0f}"},
        {"name": "H2H",      "agrees": bool(h2h_avg is not None and ((is_over and h2h_avg>line) or (not is_over and h2h_avg<line))), "detail": f"{h2h_avg:.1f}" if h2h_avg else "N/A"},
        {"name": "Pace",     "agrees": (is_over and f.get("pace_rank",15)>15) or (not is_over and f.get("pace_rank",15)<15), "detail": f"#{f.get('pace_rank',15):.0f}"},
        {"name": "Windows",  "agrees": bool(f.get("all_windows_over") or f.get("all_windows_under")), "detail": "aligned"},
        {"name": "Rest",     "agrees": not bool(f.get("is_b2b", False)), "detail": f"{int(f.get('rest_days',2))}d"},
    ]
    return [{"name": d["name"], "agrees": bool(d["agrees"]), "detail": str(d["detail"])} for d in details]

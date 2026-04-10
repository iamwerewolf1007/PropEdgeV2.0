"""
PropEdge V2.0 — generate_dataset.py
────────────────────────────────────────────────────────────────────────────────
Builds the enriched ML dataset for all graded plays across both seasons.

Source files expected in raw/:
  Player_lines_2024-25.xlsx
  Player_lines_2025-26.xlsx
  gamelogs_2024_25.csv
  gamelogs_2025_26.csv
  h2h_database.csv

Outputs:
  output/propedge_v2_ml_dataset.csv    ← ML pipeline input
  output/propedge_v2_ml_dataset.xlsx   ← human review (3 sheets)
  output/propedge_v2_schema.csv        ← column documentation
  output/propedge_features.csv         ← feature cache for train.py

RUN: python3 run.py generate-dataset
────────────────────────────────────────────────────────────────────────────────
"""
from __future__ import annotations

import warnings
from pathlib import Path

import numpy as np
import pandas as pd

from config import (
    RAW_DATA_DIR, OUTPUT_DIR, FEATURES_CSV, DATASET_CSV, DATASET_XLSX,
    SCHEMA_CSV, MODELS_DIR, MODEL_FILE,
    PROP_LINES_2425, PROP_LINES_2526,
    GAMELOGS_2425_CSV, GAMELOGS_2526_CSV,
    H2H_FILE, VERSION_TAG,
)
from feature_engine import build_features, MODEL_FEATURES
from player_name_aliases import build_nmap, resolve_name
from reasoning_engine import generate_pre_match_reason, generate_post_match_reason

warnings.filterwarnings("ignore")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ── Loaders ───────────────────────────────────────────────────────────────────

def load_prop_lines() -> pd.DataFrame:
    dfs = []
    for fname, season in [
        (PROP_LINES_2425, "2024-25"),
        (PROP_LINES_2526, "2025-26"),
    ]:
        path = RAW_DATA_DIR / fname
        if not path.exists():
            raise FileNotFoundError(
                f"Missing: {path}\n"
                f"Place source files in {RAW_DATA_DIR}/\n"
                f"Expected filename: {fname}"
            )
        df = pd.read_excel(path)
        df["season"] = season
        dfs.append(df)
    out = pd.concat(dfs, ignore_index=True)
    out["Date"] = pd.to_datetime(out["Date"])
    out.rename(columns={
        "Date": "game_date", "Player": "player_name", "Position": "position",
        "Line": "prop_line", "Over Odds": "over_odds", "Under Odds": "under_odds",
        "Books": "books", "Min Line": "min_line", "Max Line": "max_line",
        "Home": "home_team", "Away": "away_team", "Game": "matchup",
    }, inplace=True)
    return out


def load_gamelogs() -> pd.DataFrame:
    dfs = []
    for fname in [GAMELOGS_2425_CSV, GAMELOGS_2526_CSV]:
        path = RAW_DATA_DIR / fname
        if not path.exists():
            raise FileNotFoundError(
                f"Missing: {path}\n"
                f"Place game log CSV files in {RAW_DATA_DIR}/\n"
                f"Expected filenames: {GAMELOGS_2425_CSV}  and  {GAMELOGS_2526_CSV}"
            )
        dfs.append(pd.read_csv(path, low_memory=False))
    out = pd.concat(dfs, ignore_index=True)
    out["GAME_DATE"] = pd.to_datetime(out["GAME_DATE"], format="mixed")
    return out


def load_h2h() -> pd.DataFrame:
    path = RAW_DATA_DIR / H2H_FILE
    if not path.exists():
        raise FileNotFoundError(
            f"Missing: {path}\n"
            f"Place H2H file in {RAW_DATA_DIR}/\n"
            f"Expected filename: {H2H_FILE}"
        )
    return pd.read_csv(path)


# ── Opponent B2B ──────────────────────────────────────────────────────────────

def build_opp_b2b(gamelogs: pd.DataFrame) -> pd.DataFrame:
    td = (gamelogs[["GAME_TEAM_ABBREVIATION", "GAME_DATE"]]
          .drop_duplicates()
          .sort_values(["GAME_TEAM_ABBREVIATION", "GAME_DATE"]))
    td["prev"]    = td.groupby("GAME_TEAM_ABBREVIATION")["GAME_DATE"].shift(1)
    td["opp_b2b"] = ((td["GAME_DATE"] - td["prev"]).dt.days == 1).astype(int)
    return td[["GAME_TEAM_ABBREVIATION", "GAME_DATE", "opp_b2b"]].rename(
        columns={"GAME_TEAM_ABBREVIATION": "opponent", "GAME_DATE": "game_date"}
    )


# ── Game log column map ───────────────────────────────────────────────────────

_GL_RENAME = {
    "PLAYER_NAME": "player_name", "GAME_DATE": "game_date",
    "PTS": "actual_pts", "MIN_NUM": "actual_min",
    "FGA": "actual_fga", "FGM": "actual_fgm",
    "FTA": "actual_fta", "FTM": "actual_ftm",
    "USAGE_APPROX": "usage_actual",
    "IS_HOME": "is_home", "OPPONENT": "opponent",
    "GAME_TEAM_ABBREVIATION": "team_abbr", "PLAYER_POSITION": "position_gl",
    "TRUE_SHOOTING_PCT": "ts_pct", "WL_WIN": "wl_win", "PLUS_MINUS": "plus_minus",
    "L3_PTS": "L3",   "L5_PTS": "L5",   "L10_PTS": "L10",
    "L20_PTS": "L20", "L30_PTS": "L30", "L50_PTS": "L50",
    "L100_PTS": "L100", "L200_PTS": "L200",
    "L3_MIN_NUM": "min_l3",   "L5_MIN_NUM": "min_l5",
    "L10_MIN_NUM": "min_l10", "L30_MIN_NUM": "min_l30",
    "L3_FGA": "fga_l3",   "L5_FGA": "fga_l5",
    "L10_FGA": "fga_l10", "L30_FGA": "fga_l30",
    "L3_FTA": "fta_l3",   "L5_FTA": "fta_l5",   "L10_FTA": "fta_l10",
    "L3_FG_PCT": "fg_pct_l3",   "L5_FG_PCT": "fg_pct_l5",
    "L10_FG_PCT": "fg_pct_l10", "L30_FG_PCT": "fg_pct_l30",
    "L3_TRUE_SHOOTING_PCT": "ts_l3",   "L5_TRUE_SHOOTING_PCT": "ts_l5",
    "L10_TRUE_SHOOTING_PCT": "ts_l10",
    "L3_USAGE_APPROX": "usage_l3",   "L5_USAGE_APPROX": "usage_l5",
    "L10_USAGE_APPROX": "usage_l10", "L30_USAGE_APPROX": "usage_l30",
    "L3_PLUS_MINUS": "pm_l3",   "L5_PLUS_MINUS": "pm_l5",
    "L10_PLUS_MINUS": "pm_l10", "L30_PLUS_MINUS": "pm_l30",
    "L3_WL_WIN": "wl_l3",  "L5_WL_WIN": "wl_l5",  "L10_WL_WIN": "wl_l10",
    "L3_IS_HOME": "home_l3", "L5_IS_HOME": "home_l5", "L10_IS_HOME": "home_l10",
    "L3_EFF_FG_PCT": "efg_l3", "L5_EFF_FG_PCT": "efg_l5", "L10_EFF_FG_PCT": "efg_l10",
    "L3_FGM": "fgm_l3", "L5_FGM": "fgm_l5",
    "L3_FTM": "ftm_l3", "L5_FTM": "ftm_l5",
}

_H2H_KEEP = [
    "PLAYER_NAME", "OPPONENT",
    "H2H_GAMES", "H2H_AVG_PTS", "H2H_STD_PTS",
    "H2H_CONFIDENCE", "H2H_PREDICTABILITY", "H2H_RECENCY_WEIGHT",
    "H2H_HOME_AVG_PTS", "H2H_AWAY_AVG_PTS", "H2H_HOME_AWAY_DIFF",
    "L3_H2H_AVG_PTS", "L5_H2H_AVG_PTS", "H2H_PTS_TREND",
    "H2H_FGA_VS_OVERALL", "H2H_TS_VS_OVERALL", "H2H_MIN_VS_OVERALL",
    "H2H_USAGE_VS_OVERALL", "H2H_WIN_AVG_PTS", "H2H_LOSS_AVG_PTS",
    "H2H_SEASON_DRIFT", "H2H_CURRENT_SEASON_AVG_PTS",
    "H2H_MEDIAN_PTS", "H2H_MAX_PTS", "H2H_MIN_PTS",
    "H2H_WINS", "H2H_LOSSES",
    "H2H_AVG_MIN", "H2H_AVG_FGA", "H2H_AVG_FTA",
    "H2H_FG_PCT", "H2H_TS_PCT",
    "H2H_SCORING_PROFILE",
]


def slim_gamelogs(gamelogs: pd.DataFrame) -> pd.DataFrame:
    keep = [c for c in _GL_RENAME if c in gamelogs.columns]
    gl   = gamelogs[keep].copy()
    gl.rename(columns={k: v for k, v in _GL_RENAME.items() if k in gl.columns}, inplace=True)
    return gl


# ── Model outputs ─────────────────────────────────────────────────────────────

def add_model_outputs(df: pd.DataFrame) -> pd.DataFrame:
    if not MODEL_FILE.exists():
        print("  ⚠ Model not found — model output columns will be null.")
        print("    Run: python3 run.py train  to generate model outputs.")
        for col in ["p_over", "p_under", "confidence_pct",
                    "predicted_pts", "market_edge_over", "market_edge_under"]:
            df[col] = np.nan
        df["direction"]  = ""
        df["tier"]       = "UNSCORED"
        df["value_flag"] = ""
        return df
    from model_engine import predict
    return predict(df)


# ── Outcome columns ───────────────────────────────────────────────────────────

def add_outcomes(df: pd.DataFrame) -> pd.DataFrame:
    df["delta"]        = df["actual_pts"] - df["prop_line"]
    df["result"]       = np.where(df["actual_pts"] > df["prop_line"], "WIN",
                         np.where(df["actual_pts"] < df["prop_line"], "LOSS", "PUSH"))
    df["result_over"]  = (df["actual_pts"] > df["prop_line"]).astype(int)
    df["result_under"] = (df["actual_pts"] < df["prop_line"]).astype(int)
    df["is_push"]      = (df["actual_pts"] == df["prop_line"]).astype(int)
    df["correct"]      = (df["direction"] == df["direction_actual"]).astype(int)
    return df


# ── Narratives ────────────────────────────────────────────────────────────────

def add_narratives(df: pd.DataFrame) -> pd.DataFrame:
    pre_list  = []
    post_list = []
    loss_list = []
    for _, row in df.iterrows():
        play = row.to_dict()
        pre  = generate_pre_match_reason(play)
        post, ltype = generate_post_match_reason(
            play,
            box_data={
                "actual_pts": row.get("actual_pts"),
                "actual_min": row.get("actual_min"),
                "actual_fga": row.get("actual_fga"),
                "actual_fgm": row.get("actual_fgm"),
            },
        )
        pre_list.append(pre)
        post_list.append(post)
        loss_list.append(ltype)
    df["pre_match_reason"]  = pre_list
    df["post_match_reason"] = post_list
    df["loss_type"]         = loss_list
    return df


# ── Schema ────────────────────────────────────────────────────────────────────

SCHEMA: list[dict] = [
    # Group A — Identity
    {"Column":"game_date",          "Group":"A — Identity",   "Description":"Game date (YYYY-MM-DD)",                                    "Type":"date",   "LeakageRisk":"No"},
    {"Column":"player_name",        "Group":"A — Identity",   "Description":"Player full name (canonical NBA API format)",               "Type":"string", "LeakageRisk":"No"},
    {"Column":"position",           "Group":"A — Identity",   "Description":"Player position (Guard/Forward/Center/hybrid)",             "Type":"string", "LeakageRisk":"No"},
    {"Column":"team_abbr",          "Group":"A — Identity",   "Description":"Player team abbreviation",                                  "Type":"string", "LeakageRisk":"No"},
    {"Column":"opponent",           "Group":"A — Identity",   "Description":"Opponent team abbreviation",                                "Type":"string", "LeakageRisk":"No"},
    {"Column":"matchup",            "Group":"A — Identity",   "Description":"Full game matchup string",                                  "Type":"string", "LeakageRisk":"No"},
    {"Column":"is_home",            "Group":"A — Identity",   "Description":"1 if player's team is home, 0 if away",                    "Type":"int",    "LeakageRisk":"No"},
    {"Column":"season",             "Group":"A — Identity",   "Description":"NBA season (2024-25 or 2025-26)",                          "Type":"string", "LeakageRisk":"No"},
    {"Column":"home_team",          "Group":"A — Identity",   "Description":"Home team name",                                           "Type":"string", "LeakageRisk":"No"},
    {"Column":"away_team",          "Group":"A — Identity",   "Description":"Away team name",                                           "Type":"string", "LeakageRisk":"No"},
    # Group B — Bookmaker
    {"Column":"prop_line",          "Group":"B — Bookmaker",  "Description":"Points prop line set by bookmakers",                       "Type":"float",  "LeakageRisk":"No"},
    {"Column":"over_odds",          "Group":"B — Bookmaker",  "Description":"American odds for OVER (e.g. -110, +120)",                 "Type":"int",    "LeakageRisk":"No"},
    {"Column":"under_odds",         "Group":"B — Bookmaker",  "Description":"American odds for UNDER",                                  "Type":"int",    "LeakageRisk":"No"},
    {"Column":"books",              "Group":"B — Bookmaker",  "Description":"Number of bookmakers setting this line",                   "Type":"int",    "LeakageRisk":"No"},
    {"Column":"min_line",           "Group":"B — Bookmaker",  "Description":"Lowest line across all books",                            "Type":"float",  "LeakageRisk":"No"},
    {"Column":"max_line",           "Group":"B — Bookmaker",  "Description":"Highest line across all books",                           "Type":"float",  "LeakageRisk":"No"},
    {"Column":"line_spread",        "Group":"B — Bookmaker",  "Description":"Max line − min line (market disagreement)",               "Type":"float",  "LeakageRisk":"No"},
    {"Column":"implied_over_prob",  "Group":"B — Bookmaker",  "Description":"Implied P(OVER) from over_odds",                          "Type":"float",  "LeakageRisk":"No"},
    {"Column":"implied_under_prob", "Group":"B — Bookmaker",  "Description":"Implied P(UNDER) from under_odds",                        "Type":"float",  "LeakageRisk":"No"},
    # Group C — Rolling
    {"Column":"L3",                 "Group":"C — Rolling",    "Description":"Avg points last 3 games (pre-game)",                      "Type":"float",  "LeakageRisk":"No"},
    {"Column":"L5",                 "Group":"C — Rolling",    "Description":"Avg points last 5 games (pre-game)",                      "Type":"float",  "LeakageRisk":"No"},
    {"Column":"L10",                "Group":"C — Rolling",    "Description":"Avg points last 10 games (pre-game)",                     "Type":"float",  "LeakageRisk":"No"},
    {"Column":"L20",                "Group":"C — Rolling",    "Description":"Avg points last 20 games",                               "Type":"float",  "LeakageRisk":"No"},
    {"Column":"L30",                "Group":"C — Rolling",    "Description":"Avg points last 30 games",                               "Type":"float",  "LeakageRisk":"No"},
    {"Column":"L50",                "Group":"C — Rolling",    "Description":"Avg points last 50 games",                               "Type":"float",  "LeakageRisk":"No"},
    {"Column":"L100",               "Group":"C — Rolling",    "Description":"Avg points last 100 games",                              "Type":"float",  "LeakageRisk":"No"},
    {"Column":"L200",               "Group":"C — Rolling",    "Description":"Avg points last 200 games (career baseline)",            "Type":"float",  "LeakageRisk":"No"},
    {"Column":"std10",              "Group":"C — Rolling",    "Description":"Approx std deviation last 10 games (|L10-L5|×1.4)",      "Type":"float",  "LeakageRisk":"No"},
    {"Column":"min_l10",            "Group":"C — Rolling",    "Description":"Avg minutes last 10 games",                              "Type":"float",  "LeakageRisk":"No"},
    {"Column":"min_l30",            "Group":"C — Rolling",    "Description":"Avg minutes last 30 games",                              "Type":"float",  "LeakageRisk":"No"},
    {"Column":"fga_l10",            "Group":"C — Rolling",    "Description":"Avg field goal attempts last 10 games",                  "Type":"float",  "LeakageRisk":"No"},
    {"Column":"fga_l30",            "Group":"C — Rolling",    "Description":"Avg field goal attempts last 30 games",                  "Type":"float",  "LeakageRisk":"No"},
    {"Column":"fta_l10",            "Group":"C — Rolling",    "Description":"Avg free throw attempts last 10 games",                  "Type":"float",  "LeakageRisk":"No"},
    {"Column":"fg_pct_l10",         "Group":"C — Rolling",    "Description":"FG% last 10 games",                                     "Type":"float",  "LeakageRisk":"No"},
    {"Column":"fg_pct_l30",         "Group":"C — Rolling",    "Description":"FG% last 30 games",                                     "Type":"float",  "LeakageRisk":"No"},
    {"Column":"ts_l3",              "Group":"C — Rolling",    "Description":"True shooting % last 3 games",                          "Type":"float",  "LeakageRisk":"No"},
    {"Column":"ts_l5",              "Group":"C — Rolling",    "Description":"True shooting % last 5 games",                          "Type":"float",  "LeakageRisk":"No"},
    {"Column":"ts_l10",             "Group":"C — Rolling",    "Description":"True shooting % last 10 games",                         "Type":"float",  "LeakageRisk":"No"},
    {"Column":"usage_l10",          "Group":"C — Rolling",    "Description":"Usage rate last 10 games",                              "Type":"float",  "LeakageRisk":"No"},
    {"Column":"usage_l30",          "Group":"C — Rolling",    "Description":"Usage rate last 30 games",                              "Type":"float",  "LeakageRisk":"No"},
    {"Column":"pm_l3",              "Group":"C — Rolling",    "Description":"Avg plus/minus last 3 games",                           "Type":"float",  "LeakageRisk":"No"},
    {"Column":"pm_l10",             "Group":"C — Rolling",    "Description":"Avg plus/minus last 10 games",                          "Type":"float",  "LeakageRisk":"No"},
    {"Column":"pm_l30",             "Group":"C — Rolling",    "Description":"Avg plus/minus last 30 games",                          "Type":"float",  "LeakageRisk":"No"},
    {"Column":"wl_l3",              "Group":"C — Rolling",    "Description":"Win rate last 3 games",                                 "Type":"float",  "LeakageRisk":"No"},
    {"Column":"wl_l5",              "Group":"C — Rolling",    "Description":"Win rate last 5 games",                                 "Type":"float",  "LeakageRisk":"No"},
    {"Column":"wl_l10",             "Group":"C — Rolling",    "Description":"Win rate last 10 games",                                "Type":"float",  "LeakageRisk":"No"},
    {"Column":"home_l3",            "Group":"C — Rolling",    "Description":"Home game fraction last 3 games",                       "Type":"float",  "LeakageRisk":"No"},
    {"Column":"home_l5",            "Group":"C — Rolling",    "Description":"Home game fraction last 5 games",                       "Type":"float",  "LeakageRisk":"No"},
    {"Column":"efg_l3",             "Group":"C — Rolling",    "Description":"Effective FG% last 3 games",                            "Type":"float",  "LeakageRisk":"No"},
    {"Column":"efg_l5",             "Group":"C — Rolling",    "Description":"Effective FG% last 5 games",                            "Type":"float",  "LeakageRisk":"No"},
    {"Column":"efg_l10",            "Group":"C — Rolling",    "Description":"Effective FG% last 10 games",                           "Type":"float",  "LeakageRisk":"No"},
    # Group D — Form
    {"Column":"momentum",           "Group":"D — Form",       "Description":"L5 − L30 (short vs long-term trend)",                   "Type":"float",  "LeakageRisk":"No"},
    {"Column":"accel",              "Group":"D — Form",       "Description":"L3 − L5 (acceleration/deceleration)",                   "Type":"float",  "LeakageRisk":"No"},
    {"Column":"reversion",          "Group":"D — Form",       "Description":"L10 − L30 (mean reversion signal)",                     "Type":"float",  "LeakageRisk":"No"},
    {"Column":"volume",             "Group":"D — Form",       "Description":"L30 − prop_line (positive = line set below average)",   "Type":"float",  "LeakageRisk":"No"},
    {"Column":"line_zscore",        "Group":"D — Form",       "Description":"(prop_line − L10) / std10",                            "Type":"float",  "LeakageRisk":"No"},
    {"Column":"l10_vs_line",        "Group":"D — Form",       "Description":"L10 − prop_line",                                      "Type":"float",  "LeakageRisk":"No"},
    {"Column":"l5_vs_line",         "Group":"D — Form",       "Description":"L5 − prop_line",                                       "Type":"float",  "LeakageRisk":"No"},
    {"Column":"l3_vs_line",         "Group":"D — Form",       "Description":"L3 − prop_line",                                       "Type":"float",  "LeakageRisk":"No"},
    {"Column":"all_windows_over",   "Group":"D — Form",       "Description":"1 if L3, L5, L10 all > prop_line",                     "Type":"int",    "LeakageRisk":"No"},
    {"Column":"all_windows_under",  "Group":"D — Form",       "Description":"1 if L3, L5, L10 all < prop_line",                     "Type":"int",    "LeakageRisk":"No"},
    {"Column":"long_term_pts",      "Group":"D — Form",       "Description":"L100 or L200 average (career baseline)",               "Type":"float",  "LeakageRisk":"No"},
    {"Column":"long_vs_short",      "Group":"D — Form",       "Description":"long_term_pts − L10",                                  "Type":"float",  "LeakageRisk":"No"},
    # Group E — Efficiency
    {"Column":"ts_trend_l3",        "Group":"E — Efficiency", "Description":"ts_l3 − fg_pct_l10 (shooting efficiency trend)",       "Type":"float",  "LeakageRisk":"No"},
    {"Column":"fg_trend",           "Group":"E — Efficiency", "Description":"fg_pct_l3 − fg_pct_l30 (FG% trend)",                  "Type":"float",  "LeakageRisk":"No"},
    {"Column":"fga_trend_l3",       "Group":"E — Efficiency", "Description":"fga_l3 − fga_l10 (shot volume trend)",                 "Type":"float",  "LeakageRisk":"No"},
    {"Column":"usage_trend_l3",     "Group":"E — Efficiency", "Description":"usage_l3 − usage_l10 (usage rate trend)",             "Type":"float",  "LeakageRisk":"No"},
    {"Column":"pm_trend",           "Group":"E — Efficiency", "Description":"pm_l3 − pm_l10 (plus/minus momentum)",                "Type":"float",  "LeakageRisk":"No"},
    # Group F — Role
    {"Column":"starter_flag",       "Group":"F — Role",       "Description":"1.0=starter(>28min) 0.5=rotation 0=bench",            "Type":"float",  "LeakageRisk":"No"},
    {"Column":"minutes_cv_score",   "Group":"F — Role",       "Description":"std10 / L10 (minutes consistency)",                   "Type":"float",  "LeakageRisk":"No"},
    {"Column":"efficiency_l5",      "Group":"F — Role",       "Description":"pts/min L5 vs pts/min L30 (efficiency trend)",        "Type":"float",  "LeakageRisk":"No"},
    {"Column":"pos_code",           "Group":"F — Role",       "Description":"0=Guard 1=Forward 2=Center",                          "Type":"float",  "LeakageRisk":"No"},
    {"Column":"pos_line_pct",       "Group":"F — Role",       "Description":"Percentile of prop_line in position's season dist",   "Type":"float",  "LeakageRisk":"No"},
    # Group G — H2H
    {"Column":"H2H_GAMES",          "Group":"G — H2H",        "Description":"Total H2H games vs this opponent",                    "Type":"int",    "LeakageRisk":"No"},
    {"Column":"H2H_AVG_PTS",        "Group":"G — H2H",        "Description":"Historical avg pts vs this opponent",                 "Type":"float",  "LeakageRisk":"No"},
    {"Column":"H2H_HOME_AVG_PTS",   "Group":"G — H2H",        "Description":"Avg pts at home vs this opponent",                   "Type":"float",  "LeakageRisk":"No"},
    {"Column":"H2H_AWAY_AVG_PTS",   "Group":"G — H2H",        "Description":"Avg pts away vs this opponent",                      "Type":"float",  "LeakageRisk":"No"},
    {"Column":"H2H_CONFIDENCE",     "Group":"G — H2H",        "Description":"Confidence weight in H2H data (0-1)",                 "Type":"float",  "LeakageRisk":"No"},
    {"Column":"H2H_PREDICTABILITY", "Group":"G — H2H",        "Description":"Historical outcome predictability (0-1)",             "Type":"float",  "LeakageRisk":"No"},
    {"Column":"H2H_PTS_TREND",      "Group":"G — H2H",        "Description":"H2H scoring trend vs this opponent",                  "Type":"float",  "LeakageRisk":"No"},
    {"Column":"H2H_SEASON_DRIFT",   "Group":"G — H2H",        "Description":"Current season avg vs historical avg vs this opp",   "Type":"float",  "LeakageRisk":"No"},
    {"Column":"h2h_gap",            "Group":"G — H2H",        "Description":"H2H_AVG_PTS − prop_line (positive = OVER edge)",     "Type":"float",  "LeakageRisk":"No"},
    {"Column":"h2h_l3_gap",         "Group":"G — H2H",        "Description":"L3_H2H_AVG_PTS − prop_line",                         "Type":"float",  "LeakageRisk":"No"},
    {"Column":"h2h_gap_abs",        "Group":"G — H2H",        "Description":"abs(h2h_gap) — magnitude of H2H edge",               "Type":"float",  "LeakageRisk":"No"},
    {"Column":"h2h_home_away_edge", "Group":"G — H2H",        "Description":"Venue-specific H2H avg − line (strongest V2 signal)","Type":"float",  "LeakageRisk":"No"},
    {"Column":"h2h_win_loss_gap",   "Group":"G — H2H",        "Description":"H2H win avg pts − H2H loss avg pts",                 "Type":"float",  "LeakageRisk":"No"},
    {"Column":"h2h_valid",          "Group":"G — H2H",        "Description":"1 if H2H_GAMES >= 5",                                "Type":"int",    "LeakageRisk":"No"},
    {"Column":"h2h_l10_agree",      "Group":"G — H2H",        "Description":"1 if H2H gap and L10 agree on direction",            "Type":"int",    "LeakageRisk":"No"},
    {"Column":"h2h_all_agree",      "Group":"G — H2H",        "Description":"1 if H2H, L10, L5 all point OVER — triple lock",    "Type":"int",    "LeakageRisk":"No"},
    {"Column":"h2h_all_under",      "Group":"G — H2H",        "Description":"1 if H2H, L10, L5 all point UNDER — triple lock",   "Type":"int",    "LeakageRisk":"No"},
    {"Column":"h2h_ts_delta",       "Group":"G — H2H",        "Description":"H2H true shooting % vs overall TS%",                "Type":"float",  "LeakageRisk":"No"},
    {"Column":"h2h_fga_delta",      "Group":"G — H2H",        "Description":"H2H shot volume vs overall FGA",                    "Type":"float",  "LeakageRisk":"No"},
    {"Column":"h2h_min_delta",      "Group":"G — H2H",        "Description":"H2H minutes vs overall minutes",                    "Type":"float",  "LeakageRisk":"No"},
    # Group H — Market
    {"Column":"line_sharpness",     "Group":"H — Market",     "Description":"min_line / max_line (1.0 = full consensus)",         "Type":"float",  "LeakageRisk":"No"},
    {"Column":"books_log",          "Group":"H — Market",     "Description":"log(books + 1) — liquidity weight",                  "Type":"float",  "LeakageRisk":"No"},
    {"Column":"line_movement_norm", "Group":"H — Market",     "Description":"line_spread / L30 (normalised market movement)",     "Type":"float",  "LeakageRisk":"No"},
    # Group I — Context
    {"Column":"opp_b2b",            "Group":"I — Context",    "Description":"1 if opponent is on back-to-back",                   "Type":"int",    "LeakageRisk":"No"},
    {"Column":"day_of_week",        "Group":"I — Context",    "Description":"Day of week (0=Monday, 6=Sunday)",                   "Type":"int",    "LeakageRisk":"No"},
    {"Column":"season_pct",         "Group":"I — Context",    "Description":"Season progress 0.0=opening, 1.0=final game",        "Type":"float",  "LeakageRisk":"No"},
    {"Column":"season_pct_sq",      "Group":"I — Context",    "Description":"season_pct² (non-linear load management effect)",    "Type":"float",  "LeakageRisk":"No"},
    {"Column":"season_segment",     "Group":"I — Context",    "Description":"Season bucket 0-4 (0=opening, 4=closing)",           "Type":"float",  "LeakageRisk":"No"},
    {"Column":"post_allstar_flag",  "Group":"I — Context",    "Description":"1 after All-Star break (Feb 18+)",                   "Type":"int",    "LeakageRisk":"No"},
    # Group J — Player signals
    {"Column":"hot_streak",              "Group":"J — Signals","Description":"Consecutive OVERs BEFORE this game (no leakage)",    "Type":"int",   "LeakageRisk":"No"},
    {"Column":"cold_streak",             "Group":"J — Signals","Description":"Consecutive UNDERs BEFORE this game (no leakage)",   "Type":"int",   "LeakageRisk":"No"},
    {"Column":"personal_over_rate",      "Group":"J — Signals","Description":"Expanding % player beats this line bucket (pre-game)","Type":"float","LeakageRisk":"No"},
    {"Column":"personal_under_rate",     "Group":"J — Signals","Description":"Expanding % player falls under this line bucket",    "Type":"float", "LeakageRisk":"No"},
    {"Column":"recency_weighted_hr_over","Group":"J — Signals","Description":"Recency-weighted OVER rate (L3×3, L4-7×2, L8-10×1)","Type":"float","LeakageRisk":"No"},
    {"Column":"recency_weighted_hr_under","Group":"J — Signals","Description":"Recency-weighted UNDER rate",                      "Type":"float", "LeakageRisk":"No"},
    # Group K — Model outputs
    {"Column":"p_over",            "Group":"K — Model",      "Description":"Calibrated P(player scores OVER line)",               "Type":"float",  "LeakageRisk":"No"},
    {"Column":"p_under",           "Group":"K — Model",      "Description":"Calibrated P(UNDER) = 1 − p_over",                  "Type":"float",  "LeakageRisk":"No"},
    {"Column":"direction",         "Group":"K — Model",      "Description":"Predicted direction: OVER or UNDER",                  "Type":"string", "LeakageRisk":"No"},
    {"Column":"confidence_pct",    "Group":"K — Model",      "Description":"Direction-specific confidence: max(p_over,p_under)×100","Type":"float","LeakageRisk":"No"},
    {"Column":"tier",              "Group":"K — Model",      "Description":"APEX/ULTRA/ELITE/STRONG/PLAY/SKIP",                  "Type":"string", "LeakageRisk":"No"},
    {"Column":"predicted_pts",     "Group":"K — Model",      "Description":"Model predicted points",                             "Type":"float",  "LeakageRisk":"No"},
    {"Column":"value_flag",        "Group":"K — Model",      "Description":"strong_over/mod_over/mod_under/strong_under",        "Type":"string", "LeakageRisk":"No"},
    {"Column":"market_edge_over",  "Group":"K — Model",      "Description":"p_over − implied_over_prob",                        "Type":"float",  "LeakageRisk":"No"},
    {"Column":"market_edge_under", "Group":"K — Model",      "Description":"p_under − implied_under_prob",                      "Type":"float",  "LeakageRisk":"No"},
    # Group L — Post-game outcomes (NEVER model inputs)
    {"Column":"actual_pts",        "Group":"L — Outcome",    "Description":"Actual points scored — POST-GAME. Never a model input.","Type":"float","LeakageRisk":"YES"},
    {"Column":"actual_min",        "Group":"L — Outcome",    "Description":"Actual minutes played — POST-GAME",                  "Type":"float",  "LeakageRisk":"YES"},
    {"Column":"actual_fga",        "Group":"L — Outcome",    "Description":"Actual FGA — POST-GAME",                            "Type":"float",  "LeakageRisk":"YES"},
    {"Column":"actual_fgm",        "Group":"L — Outcome",    "Description":"Actual FGM — POST-GAME",                            "Type":"float",  "LeakageRisk":"YES"},
    {"Column":"actual_fta",        "Group":"L — Outcome",    "Description":"Actual FTA — POST-GAME",                            "Type":"float",  "LeakageRisk":"YES"},
    {"Column":"actual_ftm",        "Group":"L — Outcome",    "Description":"Actual FTM — POST-GAME",                            "Type":"float",  "LeakageRisk":"YES"},
    {"Column":"delta",             "Group":"L — Outcome",    "Description":"actual_pts − prop_line (positive=over)",             "Type":"float",  "LeakageRisk":"YES"},
    {"Column":"result",            "Group":"L — Outcome",    "Description":"WIN/LOSS/PUSH",                                      "Type":"string", "LeakageRisk":"YES"},
    {"Column":"result_over",       "Group":"L — Outcome",    "Description":"1 if actual_pts > prop_line",                       "Type":"int",    "LeakageRisk":"YES"},
    {"Column":"result_under",      "Group":"L — Outcome",    "Description":"1 if actual_pts < prop_line",                       "Type":"int",    "LeakageRisk":"YES"},
    {"Column":"is_push",           "Group":"L — Outcome",    "Description":"1 if actual_pts == prop_line",                      "Type":"int",    "LeakageRisk":"YES"},
    {"Column":"direction_actual",  "Group":"L — Outcome",    "Description":"OVER or UNDER (actual outcome)",                    "Type":"string", "LeakageRisk":"YES"},
    {"Column":"correct",           "Group":"L — Outcome",    "Description":"1 if model direction == actual direction",           "Type":"int",    "LeakageRisk":"YES"},
    {"Column":"loss_type",         "Group":"L — Outcome",    "Description":"CLOSE_CALL/MINUTES_SHORTFALL/SHOOTING_VARIANCE/BLOWOUT_EFFECT/TREND_REVERSAL/H2H_OVERRIDE_FAILURE/HIGH_CONF_ANOMALY/MODEL_FAILURE_GENERAL/MODEL_CORRECT","Type":"string","LeakageRisk":"YES"},
    # Group M — Narratives
    {"Column":"pre_match_reason",  "Group":"M — Narrative",  "Description":"6-part pre-match plain-English reasoning",           "Type":"string", "LeakageRisk":"No"},
    {"Column":"post_match_reason", "Group":"M — Narrative",  "Description":"7-part post-match narrative with loss classification","Type":"string", "LeakageRisk":"YES"},
]


def write_schema() -> None:
    pd.DataFrame(SCHEMA).to_csv(SCHEMA_CSV, index=False)
    print(f"  ✓ Schema → {SCHEMA_CSV}  ({len(SCHEMA)} columns documented)")


def write_xlsx(df: pd.DataFrame) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠ openpyxl not installed — skipping XLSX.")
        print("    Run: pip3 install openpyxl")
        return

    wb  = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(top=thin, left=thin, right=thin, bottom=thin)
    hdr_fill  = PatternFill("solid", fgColor="1B4F8C")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Arial", size=10)
    alt_fill  = PatternFill("solid", fgColor="EBF3FB")
    grn_fill  = PatternFill("solid", fgColor="D4EFDF")
    red_fill  = PatternFill("solid", fgColor="FDECEA")
    amb_fill  = PatternFill("solid", fgColor="FFF3CD")

    def hdr_row(ws, r, n):
        for c in range(1, n+1):
            cell = ws.cell(r, c)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.border = bdr
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def body_row(ws, r, n, alt=False):
        for c in range(1, n+1):
            cell = ws.cell(r, c)
            cell.font = body_font; cell.border = bdr
            if alt: cell.fill = alt_fill

    # Sheet 1 — Summary
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "PropEdge V2.0 — ML Dataset"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1B4F8C")
    rows = [
        ["Metric", "Value"],
        ["Total rows", len(df)],
        ["Seasons", ", ".join(df["season"].unique().tolist())],
        ["Date range", f"{df['game_date'].min()} → {df['game_date'].max()}"],
        ["Unique players", df["player_name"].nunique()],
        ["Pre-game features", len([s for s in SCHEMA if s["LeakageRisk"]=="No"])],
        ["Post-game outcomes", len([s for s in SCHEMA if s["LeakageRisk"]=="YES"])],
        ["Total documented cols", len(SCHEMA)],
        ["APEX plays",   int((df["tier"]=="APEX").sum())   if "tier" in df.columns else "N/A"],
        ["ULTRA plays",  int((df["tier"]=="ULTRA").sum())  if "tier" in df.columns else "N/A"],
        ["ELITE plays",  int((df["tier"]=="ELITE").sum())  if "tier" in df.columns else "N/A"],
        ["Baseline WIN rate", f"{df['result_over'].mean()*100:.1f}%" if "result_over" in df.columns else "N/A"],
    ]
    for i, row_data in enumerate(rows, start=3):
        for j, val in enumerate(row_data, start=1):
            cell = ws.cell(i, j, value=val)
            cell.font  = hdr_font  if i==3 else body_font
            cell.fill  = hdr_fill  if i==3 else (alt_fill if i%2==0 else PatternFill())
            cell.border = bdr
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30

    # Sheet 2 — Plays
    ws2 = wb.create_sheet("Plays")
    cols = [s["Column"] for s in SCHEMA if s["Column"] in df.columns]
    ws2.append(cols); hdr_row(ws2, 1, len(cols))
    for ri, (_, row_df) in enumerate(df[cols].iterrows(), start=2):
        ws2.append([str(v) if isinstance(v,(dict,list)) else
                    (v.isoformat() if hasattr(v,"isoformat") else v)
                    for v in row_df.values])
        rv = str(row_df.get("result",""))
        if rv=="WIN":
            for c in range(1,len(cols)+1): ws2.cell(ri,c).fill = grn_fill
        elif rv=="LOSS":
            for c in range(1,len(cols)+1): ws2.cell(ri,c).fill = red_fill
        else:
            body_row(ws2, ri, len(cols), ri%2==0)
    for i,_ in enumerate(cols,1):
        ws2.column_dimensions[get_column_letter(i)].width = 14
    ws2.freeze_panes = "A2"

    # Sheet 3 — Schema
    ws3 = wb.create_sheet("Schema")
    sc  = ["Column","Group","Description","Type","LeakageRisk"]
    ws3.append(sc); hdr_row(ws3, 1, len(sc))
    for ri, entry in enumerate(SCHEMA, start=2):
        ws3.append([entry.get(c,"") for c in sc])
        body_row(ws3, ri, len(sc), ri%2==0)
        if entry.get("LeakageRisk")=="YES":
            for c in range(1,len(sc)+1): ws3.cell(ri,c).fill = amb_fill
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 60
    ws3.column_dimensions["D"].width = 10
    ws3.column_dimensions["E"].width = 14
    ws3.freeze_panes = "A2"

    wb.save(DATASET_XLSX)
    print(f"  ✓ XLSX  → {DATASET_XLSX}  ({len(df):,} rows, {len(cols)} cols, 3 sheets)")


# ── Main ──────────────────────────────────────────────────────────────────────

def generate_dataset() -> None:
    print(f"\n  {VERSION_TAG} — Dataset Generation")
    print("  " + "─" * 52)

    print("  Loading source files...")
    lines    = load_prop_lines()
    gamelogs = load_gamelogs()
    h2h      = load_h2h()
    print(f"    Lines:    {len(lines):,} rows across 2 seasons")
    print(f"    Gamelogs: {len(gamelogs):,} rows")
    print(f"    H2H:      {len(h2h):,} player-opponent pairs")

    print("  Merging game logs...")
    gl = slim_gamelogs(gamelogs)
    df = pd.merge(lines, gl, on=["player_name","game_date"], how="left")

    print("  Merging H2H database...")
    h2h_keep = list(dict.fromkeys([c for c in _H2H_KEEP if c in h2h.columns]))
    h2h_s = h2h[h2h_keep].rename(
        columns={"PLAYER_NAME":"player_name","OPPONENT":"opponent"}
    )
    df = pd.merge(df, h2h_s, on=["player_name","opponent"], how="left")

    opp_b2b = build_opp_b2b(gamelogs)
    df = pd.merge(df, opp_b2b, on=["opponent","game_date"], how="left")
    df["opp_b2b"] = df["opp_b2b"].fillna(0).astype(int)

    graded = df[df["actual_pts"].notna()].copy()
    graded = graded[graded["actual_pts"] != graded["prop_line"]].copy()
    print(f"  Graded plays (excl. pushes): {len(graded):,}")

    print("  Computing features...")
    graded = build_features(graded)
    graded["direction_actual"] = np.where(
        graded["actual_pts"] > graded["prop_line"], "OVER", "UNDER"
    )

    print("  Running model inference...")
    graded = add_model_outputs(graded)

    graded = add_outcomes(graded)

    print("  Generating narratives...")
    graded = add_narratives(graded)

    # Order columns by schema
    schema_cols = [s["Column"] for s in SCHEMA]
    ordered     = [c for c in schema_cols if c in graded.columns]
    extras      = [c for c in graded.columns if c not in ordered]
    graded      = graded[ordered + extras]

    graded.to_csv(FEATURES_CSV,  index=False)
    graded.to_csv(DATASET_CSV,   index=False)
    print(f"  ✓ Features → {FEATURES_CSV}")
    print(f"  ✓ CSV    → {DATASET_CSV}  ({len(graded):,} rows, {len(graded.columns)} cols)")

    write_schema()

    print("  Writing XLSX...")
    write_xlsx(graded)

    print()
    print(f"  {'─'*52}")
    print(f"  Dataset generation complete.")
    print(f"  Rows:    {len(graded):,}")
    n_pre  = len([s for s in SCHEMA if s["LeakageRisk"]=="No"])
    n_post = len([s for s in SCHEMA if s["LeakageRisk"]=="YES"])
    print(f"  Columns: {len(graded.columns)} total  ({n_pre} pre-game, {n_post} outcome)")
    for s in graded["season"].unique():
        sub = graded[graded["season"]==s]
        wr  = sub["result_over"].mean()*100 if "result_over" in sub.columns else 0
        print(f"    {s}: {len(sub):,} plays  |  baseline WIN rate: {wr:.1f}%")
    if "tier" in graded.columns:
        for t in ["APEX","ULTRA","ELITE","STRONG","PLAY","SKIP"]:
            n = int((graded["tier"]==t).sum())
            if n > 0:
                acc = graded[graded["tier"]==t]["correct"].mean()*100 if "correct" in graded.columns else 0
                print(f"    {t:<8} {n:6,} plays  acc={acc:.1f}%")


if __name__ == "__main__":
    generate_dataset()
